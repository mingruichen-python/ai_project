import openpyxl
def modify_cell(wb):
    ws = wb.active

    cell_pos = input("Enter the cell position (e.g., A3): ").strip()
    new_value = input("Enter the new content: ").strip()

    try:
        ws[cell_pos] = new_value
        print(f"Cell {cell_pos} has been updated to: {new_value}")
    except Exception as e:
        print("Error:", e)

def create_sheet(wb):
    if sheet_name in wb.sheetnames:
        print("Sheet already exists.")
    sheet_name = input("Enter the name of the new sheet: ").strip()
    ws = wb.create_sheet(sheet_name)
    print(f"New sheet created: {sheet_name}")

    cell_pos = input("Enter the cell position to write (e.g., A1): ").strip()
    new_value = input("Enter the content: ").strip()

    ws[cell_pos] = new_value
    print(f"Written to {sheet_name} at {cell_pos}: {new_value}")


try:
    workbook=input("name:")
    wb = openpyxl.load_workbook(workbook)

    print("1. Modify an existing cell")
    print("2. Create a new sheet and write content")
    print("Existing sheets:", wb.sheetnames)
    choice = input("Please choose an option (1 or 2): ").strip()

    if choice == "1":
        
        modify_cell(wb)

    elif choice == "2":

        create_sheet(wb):

    else:
        print("Invalid choice. Please enter 1 or 2.")


    wb.save(workbook)
    print("succesed")
except Exception as e:
        print("Error:", e)
